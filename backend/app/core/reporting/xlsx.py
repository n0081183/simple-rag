"""XLSX report export with metadata columns."""

from __future__ import annotations

from io import BytesIO

from app.core.reporting.anonymizer import anonymize_results
from app.core.verification.schemas import VerificationResult

_HEADERS = {
    "pl": [
        "ID",
        "Tytuł",
        "Wymaganie",
        "Ocena",
        "Pewność",
        "Liczba dowodów",
        "Uwagi",
        "Źródła (pliki)",
    ],
    "en": [
        "ID",
        "Title",
        "Requirement",
        "Verdict",
        "Confidence",
        "Evidence count",
        "Caveats",
        "Source files",
    ],
}


def export_xlsx(
    results: list[VerificationResult],
    language: str = "pl",
    anonymize: bool = False,
    product: str | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if anonymize:
        results = anonymize_results(results)

    wb = Workbook()
    meta = wb.active
    meta.title = "Meta"
    meta.append(["Product", product or ""])
    meta.append(["Language", language])
    meta.append(["Total requirements", len(results)])

    ws = wb.create_sheet("Requirements")
    headers = _HEADERS.get(language, _HEADERS["en"])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in results:
        title = r.requirement_text.split()[:8]
        sources = "; ".join({e.source_file for e in r.evidence if e.source_file})
        ws.append(
            [
                r.requirement_id,
                " ".join(title)[:80],
                r.requirement_text[:2000],
                r.verdict.value,
                r.confidence.value,
                len(r.evidence),
                r.caveats or "",
                sources[:500],
            ]
        )

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["C"].width = 60

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
